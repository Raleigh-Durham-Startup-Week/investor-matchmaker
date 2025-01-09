import pandas as pd
import networkx as nx
from networkx.algorithms import bipartite
from openpyxl import load_workbook

INVESTMENT_TYPE_COLS = [
    "Investment Type: Angel",
    "Investment Type: Pre-Seed",
    "Investment Type: Seed",
    "Investment Type: Series A",
    "Investment Type: Series B"
]

BIZ_TYPE_COLS = [
    "Biz Type: B2C",
    "Biz Type: B2B",
    "Biz Type: CPG",
    "Biz Type: Enterprise"
]

SECTOR_COLS = [
    "Sector: AgTech",
    "Sector: AI",
    "Sector: CleanTech/Energy",
    "Sector: Consumer Products",
    "Sector: Consumer Services",
    "Sector: Consulting",
    "Sector: Cybersecurity",
    "Sector: DeepTech/HardTech",
    "Sector: EdTech",
    "Sector: FinTech",
    "Sector: GovTech",
    "Sector: Hardware",
    "Sector: HealthTech",
    "Sector: Life Sciences/Biotech",
    "Sector: Manufacturing/Industrial",
    "Sector: Materials",
    "Sector: Network/Telecom",
    "Sector: SaaS",
    "Sector: Social",
    "Sector: Textiles"
]

COMPANY_EXCEL = "Company Data.xlsx"
INVESTOR_EXCEL = "Investor Data.xlsx"
SCHEDULE_EXCEL = "Schedule.xlsx"

def load_data():
    companies_df = pd.read_excel(COMPANY_EXCEL)
    investors_df = pd.read_excel(INVESTOR_EXCEL)
    return companies_df, investors_df

def find_valid_pairs(companies_df, investors_df):
    valid_pairs = []
    for _, inv_row in investors_df.iterrows():
        for _, co_row in companies_df.iterrows():
            match_investment_type = any(
                inv_row[col] == "Yes" and co_row[col] == "Yes"
                for col in INVESTMENT_TYPE_COLS
            )
            match_biz_type = any(
                inv_row[col] == "Yes" and co_row[col] == "Yes"
                for col in BIZ_TYPE_COLS
            )
            match_sector = any(
                inv_row[col] == "Yes" and co_row[col] == "Yes"
                for col in SECTOR_COLS
            )
            if match_investment_type and match_biz_type and match_sector:
                investor_name = inv_row["Investment Firm"]
                company_name = co_row["Startup Name"]
                valid_pairs.append((investor_name, company_name))
    return valid_pairs

def schedule_slots(valid_pairs, investors_df, companies_df):
    time_slots = [
        "8:00 AM","8:25 AM","8:50 AM","9:15 AM","9:40 AM",
        "10:05 AM","10:30 AM","10:55 AM","11:20 AM","11:45 AM"
    ]
    schedule = [[] for _ in range(len(time_slots))]
    used_pairs = set()
    investor_names = list(investors_df["Investment Firm"].unique())
    company_names = list(companies_df["Startup Name"].unique())

    for slot_idx in range(len(time_slots)):
        graph = nx.Graph()
        graph.add_nodes_from(investor_names, bipartite=0)
        graph.add_nodes_from(company_names, bipartite=1)

        for inv, co in valid_pairs:
            if (inv, co) not in used_pairs:
                graph.add_edge(inv, co)

        matching_dict = bipartite.matching.maximum_matching(
            graph, top_nodes=set(investor_names)
        )

        matched_this_slot = []
        matched_investors = set()
        matched_companies = set()
        for inv in investor_names:
            if inv in matching_dict:
                co = matching_dict[inv]
                if co not in matched_companies and inv not in matched_investors:
                    matched_this_slot.append((inv, co))
                    matched_investors.add(inv)
                    matched_companies.add(co)

        matched_this_slot = matched_this_slot[:20]
        schedule[slot_idx] = matched_this_slot

        for pair in matched_this_slot:
            used_pairs.add(pair)

    return time_slots, schedule

def write_schedule_to_excel(time_slots, schedule):
    wb = load_workbook(SCHEDULE_EXCEL)
    sheet = wb.active
    for slot_idx, pairs in enumerate(schedule):
        row_idx = slot_idx + 2
        for table_num, (inv, co) in enumerate(pairs, start=1):
            company_col = 2 + 2*(table_num-1)
            investor_col = 3 + 2*(table_num-1)
            sheet.cell(row=row_idx, column=company_col).value = co
            sheet.cell(row=row_idx, column=investor_col).value = inv
    wb.save(SCHEDULE_EXCEL)

def main():
    companies_df, investors_df = load_data()
    valid_pairs = find_valid_pairs(companies_df, investors_df)
    time_slots, schedule = schedule_slots(valid_pairs, investors_df, companies_df)
    write_schedule_to_excel(time_slots, schedule)

if __name__ == "__main__":
    main()
